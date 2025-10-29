# app.py
# ------------------------------------------------------------
# Inkoop vergelijkingstool — export mét stabiele Excel-tabellen
# - Tab 1: Oude invoer
# - Tab 2: Nieuwe invoer
# - Tab 3: Nieuwe rijen t.o.v. oud met Delay (days) > 0
# ------------------------------------------------------------

import io
import math
import re
import traceback
from typing import List, Tuple

import pandas as pd
import streamlit as st

# ---------- Helpers ----------
def _norm_key_name(name: str) -> str:
    if name is None:
        return ""
    s = str(name).replace("\xa0", " ")
    s = " ".join(s.split()).strip().lower()
    s = re.sub(r"\s*\(\s*", "(", s)
    s = re.sub(r"\s*\)\s*", ")", s)
    return s

def _canon_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliseer kolomnamen. 'Number' losse kolom; 'Item number' is key; 'Delay (days)' voor filter."""
    rename_map = {}
    for col in df.columns:
        k = _norm_key_name(col)
        if k in {"item number", "itemnumber", "item no", "item_no", "itemnr", "item"}:
            rename_map[col] = "Item number"
        elif k in {"delay(days)", "delay (days)", "delay_days", "delay", "vertraging(dagen)", "vertraging (dagen)"}:
            rename_map[col] = "Delay (days)"
        elif k == "number":
            rename_map[col] = "Number"
    return df.rename(columns=rename_map)

def _ensure_item_number(df: pd.DataFrame) -> pd.DataFrame:
    """Fallback: als 'Item number' ontbreekt maar 'Number' bestaat, kopieer die naar 'Item number'."""
    if "Item number" not in df.columns and "Number" in df.columns:
        df = df.copy()
        df["Item number"] = df["Number"]
    return df

def _clean_key_value(v) -> str:
    """Key als nette string: 1001.0 -> '1001', NaN -> ''."""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def _prepare(df_old: pd.DataFrame, df_new: pd.DataFrame, key: str
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """Normaliseer, dedupliceer op key en aligneer kolommen (indexeer op key)."""
    df_old = _ensure_item_number(_canon_columns(df_old.copy()))
    df_new = _ensure_item_number(_canon_columns(df_new.copy()))

    if key not in df_old.columns or key not in df_new.columns:
        raise KeyError(
            f"Kolom '{key}' ontbreekt in één van de bestanden.\n"
            f"Oud kolommen: {list(df_old.columns)}\n"
            f"Nieuw kolommen: {list(df_new.columns)}"
        )

    df_old[key] = df_old[key].map(_clean_key_value)
    df_new[key] = df_new[key].map(_clean_key_value)

    df_old = df_old.drop_duplicates(subset=[key], keep="last")
    df_new = df_new.drop_duplicates(subset=[key], keep="last")

    common_cols = sorted(set(df_new.columns).intersection(df_old.columns))
    if key in common_cols:
        common_cols = [key] + [c for c in common_cols if c != key]

    df_old = df_old[common_cols].set_index(key)
    df_new = df_new[common_cols].set_index(key)
    return df_old, df_new, common_cols

def _new_rows_with_delay(df_old_idx: pd.DataFrame, df_new_idx: pd.DataFrame) -> pd.DataFrame:
    """Nieuwe Item numbers t.o.v. oud, daarna filter op Delay (days) > 0 (indien kolom bestaat)."""
    added_keys = sorted(set(df_new_idx.index) - set(df_old_idx.index))
    if not added_keys:
        cols = ["Item number"] + [c for c in df_new_idx.columns if c != "Item number"]
        return pd.DataFrame(columns=cols)
    df_added = df_new_idx.loc[added_keys].copy()
    if "Delay (days)" in df_added.columns:
        df_added["Delay (days)"] = pd.to_numeric(df_added["Delay (days)"], errors="coerce").fillna(0)
        df_added = df_added[df_added["Delay (days)"] > 0]
    return df_added.reset_index()

# ---------- Excel-safe kolomnamen ----------
_excel_bad_chars = re.compile(r"[\[\]\:\*\?\\/]")
def _excel_safe_headers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Maak kolomnamen Excel-table-proof:
    - niet leeg
    - unieke namen
    - geen problematische tekens
    """
    cols_out = []
    seen = set()
    for i, c in enumerate(df.columns, start=1):
        name = str(c) if c is not None else ""
        name = name.strip()
        if name == "":
            name = f"Column{i}"
        name = _excel_bad_chars.sub("_", name)
        base = name
        n = 1
        while name in seen:
            n += 1
            name = f"{base}_{n}"
        seen.add(name)
        cols_out.append(name)
    out = df.copy()
    out.columns = cols_out
    return out

def _safe_table_name(sheet_title: str) -> str:
    """Excel Table displayName: letters/nummers/underscore, start met letter, max ~100 chars."""
    base = re.sub(r"[^A-Za-z0-9_]", "_", sheet_title)  # vervang () > etc.
    if not base or not base[0].isalpha():
        base = "T_" + base
    base = base[:80]  # houd het kort
    from uuid import uuid4
    return f"{base}_{str(uuid4())[:8]}"

# ---------- Excel export (mét Tables, veilig) ----------
def to_excel_bytes(df_old_idx, df_new_idx, df_added_delay) -> bytes:
    from openpyxl import load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    # Alleen voor export: headers Excel-proof maken
    df_old_out = _excel_safe_headers(df_old_idx.reset_index())
    df_new_out = _excel_safe_headers(df_new_idx.reset_index())
    df_add_out = _excel_safe_headers(df_added_delay)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df_old_out.to_excel(xw, sheet_name="Oude_invoer", index=False)
        df_new_out.to_excel(xw, sheet_name="Nieuwe_invoer", index=False)
        df_add_out.to_excel(xw, sheet_name="Nieuwe_rijen_Delay_gt_0", index=False)
    buf.seek(0)

    wb = load_workbook(buf)

    def add_table(ws):
        # Excel vereist minimaal header + 1 datarij
        if ws.max_row >= 2 and ws.max_column >= 1:
            last_row = ws.max_row
            last_col_letter = get_column_letter(ws.max_column)
            ref = f"A1:{last_col_letter}{last_row}"
            tname = _safe_table_name(ws.title)
            t = Table(displayName=tname, ref=ref)
            t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            ws.add_table(t)

    for s in ["Oude_invoer", "Nieuwe_invoer", "Nieuwe_rijen_Delay_gt_0"]:
        add_table(wb[s])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

# ---------- Streamlit UI ----------
st.set_page_config(page_title="Inkoop vergelijkingstool", layout="centered")
st.title("Inkoop vergelijkingstool")
st.caption("Tab 1: oud • Tab 2: nieuw • Tab 3: nieuwe rijen t.o.v. oud met **Delay (days) > 0** — export mét Excel-tabellen")

# Uploaders ONDER ELKAAR
file_old = st.file_uploader("Oud bestand (.xlsx)", type=["xlsx"], help="Sleep hier het **oude** Excelbestand.")
file_new = st.file_uploader("Nieuw bestand (.xlsx)", type=["xlsx"], help="Sleep hier het **nieuwe** Excelbestand.")

if st.button("Vergelijk", type="primary", use_container_width=True):
    if not file_old or not file_new:
        st.warning("Kies zowel het **oude** als het **nieuwe** bestand.")
        st.stop()

    try:
        with st.spinner("Bezig met vergelijken…"):
            df_old_raw = pd.read_excel(file_old)
            df_new_raw = pd.read_excel(file_new)

            df_old_idx, df_new_idx, _ = _prepare(df_old_raw, df_new_raw, key="Item number")
            df_added_delay = _new_rows_with_delay(df_old_idx, df_new_idx)

        tabs = st.tabs(["Oude invoer", "Nieuwe invoer", "Nieuwe rijen (Delay>0)"])
        with tabs[0]:
            st.dataframe(df_old_idx.reset_index(), use_container_width=True, hide_index=True)
        with tabs[1]:
            st.dataframe(df_new_idx.reset_index(), use_container_width=True, hide_index=True)
        with tabs[2]:
            st.write(f"Nieuw t.o.v. oud met `Delay (days) > 0`: **{len(df_added_delay)}** rijen.")
            st.dataframe(df_added_delay, use_container_width=True, hide_index=True)

        excel_bytes = to_excel_bytes(df_old_idx, df_new_idx, df_added_delay)
        st.download_button(
            "📥 Download Excel (3 tabbladen, mét tabelopmaak)",
            data=excel_bytes,
            file_name="Inkoop_vergelijking.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    except Exception as e:
        st.error(f"Er is een fout opgetreden tijdens het vergelijken: {e}")
        with st.expander("Traceback"):
            st.code(traceback.format_exc())
